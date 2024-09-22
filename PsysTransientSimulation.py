'''
transient simulation
by Yao Xiao
xiaoyaodashe@outlook.com
yao.xiao.1@stonybrook.edu
09/18/2024
'''
# model from p135 in book 'POWER SYSTEM DYNAMIC AND STABILITY'

import dask
import numpy as np
import pypower.api as pp
from dask import delayed, compute
from openpyxl import load_workbook
from scipy.sparse import coo_matrix
from pypower.makeYbus import makeYbus
from pypower.idx_bus import BUS_I, PD, QD, VM, VA, GS, BUS_TYPE, PV, PQ, REF
from pypower.idx_brch import F_BUS, T_BUS, PF, PT, QF, QT
from pypower.idx_gen import PG, QG, VG, QMAX, QMIN, GEN_BUS, GEN_STATUS

# index in generator related with transience
XL, RS, XD, X_D, X__D, T_D, T__D, XQ, X_Q, X__Q, T_Q, T__Q, H, D, S1, S1_2, FRP, FRQ, TE, KE, TF, KF, AX, BX, KA, TA = (
    21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46)


def main():

    # base power and parameters
    baseMVA = 100
    omegaS = 2 * np.pi * 60
    delatT = 0.005

    # get a transient analysis object
    caseTransient = powerTransient('case118_BY_V3.xlsx', baseMVA, omegaS, delatT)

    # initialize the states variables and parameters
    caseTransient.initialization()

    # run the transient dynamic
    xSet, ySet = caseTransient.TransientSimulation(10)

    pass


class powerTransient():
    '''transient analysis of power system'''
    def __init__(self, file, baseMVA, omegaS, deltaT):
        '''get the data and run the power flow'''
        self.baseMVA = baseMVA
        self.omegaS = omegaS
        self.deltaT = deltaT
        # get the data from excel
        self.bus, self.branch, self.generator = self.impoExcel(file)
        self.nBus, self.nBranch, self.nGenerator = len(self.bus), len(self.branch), len(self.generator)
        # set dimension
        self.nDimensionx = 7*self.nGenerator
        self.nDimensiony = 2*self.nGenerator + 2*self.nBus
        self.nDimension = self.nDimensionx + self.nDimensiony
        # run the powerflow
        self.powerflowStable()
        # get result of each facility
        self.bus = self.Init_PowerFlow['bus']
        self.branch = self.Init_PowerFlow['branch']
        self.gen = self.Init_PowerFlow['gen']
        # get the bus variables and parameters in facility array
        self.busVm = self.bus[:, VM]
        self.busVa = self.bus[:, VA] * np.pi / 180
        self.busPd = self.bus[:, PD] / baseMVA
        self.busQd = self.bus[:, QD] / baseMVA
        # get the generator variables and parameters in facility array
        self.genBus = self.gen[:, 0].astype(int)
        self.genPg = self.gen[:, PG] / baseMVA
        self.genQg = self.gen[:, QG] / baseMVA
        self.genD = self.generator[:, D]
        self.genH = self.generator[:, H]
        self.genRs = self.generator[:, RS]
        self.genXq = self.generator[:, XQ]
        self.genXd = self.generator[:, XD]
        self.genX_q = self.generator[:, X_Q]
        self.genX_d = self.generator[:, X_D]
        self.genKe = self.generator[:, KE]
        self.genAx = self.generator[:, AX]
        self.genBx = self.generator[:, BX]
        self.genKf = self.generator[:, KF]
        self.genTa = self.generator[:, TA]
        self.genTf = self.generator[:, TF]
        self.genTe = self.generator[:, TE]
        self.genT_d = self.generator[:, T_D]
        self.genT_q = self.generator[:, T_Q]
        self.genKa = self.generator[:, KA]
        self.genVm = self.busVm[self.genBus - 1]
        self.genVa = self.busVa[self.genBus - 1]
        self.genOmegaS = self.omegaS * np.ones((self.nGenerator, 1))[:, 0]
        # compute M=2H/omegaS
        self.genM = 2*self.genH/self.genOmegaS
        # build admittance matrices
        bbcbus, bbcbranch = self.bus, self.branch
        bbcbus[:, BUS_I] = bbcbus[:, BUS_I] - 1
        bbcbranch[:, [F_BUS, T_BUS]] = bbcbranch[:, [F_BUS, T_BUS]] - 1
        self.Ybus, self.Yf, self.Yt = makeYbus(self.baseMVA, self.bus, self.branch)
        # build adjacent matrix(gen to bus)
        self.G2B = coo_matrix((np.ones(54), (self.genBus - 1, np.arange(0, self.nGenerator))), shape=(self.nBus, self.nGenerator)).toarray()

    def impoExcel(self, file):
        ''' get data from excel
        :param file: excel
        :return: bus, branch, generator
        '''
        # get data from excel
        PsysData = load_workbook(file)
        # specify the facility data needed
        facility = ['BUS', 'BRANCH', 'GEN']

        dataR = []
        for sheetName in facility:
            sheet = PsysData[sheetName]
            dataNp = np.array([list(row) for row in sheet.iter_rows(values_only=True)])
            dataR.append(dataNp)
        return dataR

    def powerflowStable(self):
        ''' run stable power flow as initial point of transient analysis
        :param baseMVA:
        :param bus:
        :param branch:
        :param generator:
        :return stable power flow
        '''
        bus_PowerFlow = self.bus
        branch_PowerFlow = self.branch
        gen_PowerFlow = self.generator[:, :21]
        # Create the PYPOWER case dictionary
        PSys_Set = {
            'version': '2',
            'baseMVA': self.baseMVA,
            'bus': bus_PowerFlow,
            'branch': branch_PowerFlow,
            'gen': gen_PowerFlow
        }
        # run the stable power flow
        self.Init_PowerFlow = pp.runpf(PSys_Set)[0]
        # Run power flow
        return self.Init_PowerFlow

    def initialization(self):
        ''' initialize the parameters and states variables of generator'''
        # step 1: compute delta
        self.genIGcomplex = (self.genPg - 1j * self.genQg) / (self.genVm * np.exp(-1j * self.genVa))
        self.genIG = np.abs(self.genIGcomplex)
        self.genGamma = np.angle(self.genIGcomplex)
        self.genDelta = np.angle(self.genVm * np.exp(1j * self.genVa) + (self.genRs + 1j * self.genXq) * self.genIGcomplex)
        # step 2: compute Id, Iq, Vd, Vq
        self.genIdq = self.genIG * np.exp(1j * (self.genGamma - self.genDelta + np.pi / 2))
        self.genId = self.genIdq.real
        self.genIq = self.genIdq.imag
        self.genVdq = self.genVm * np.exp(1j * (self.genVa - self.genDelta + np.pi / 2))
        self.genVd = self.genVdq.real
        self.genVq = self.genVdq.imag
        # step 3: compute E_d
        self.genE_d = self.genVd + self.genRs * self.genId - self.genX_q * self.genIq
        # step 4: compute E_q
        self.genE_q = self.genVq + self.genRs * self.genIq + self.genX_d * self.genId
        # step 5: compute Efd
        self.genEfd = self.genE_q + (self.genXd - self.genX_d) * self.genId
        # step 6: compute Rf, Vr, Vref
        self.genVr = (self.genKe + self.genAx * np.exp(self.genBx * self.genEfd)) * self.genEfd
        self.genRf = self.genKf / self.genTf * self.genEfd
        self.genVref = self.genVm + (self.genVr / self.genKa)
        # step 7: compute omega, Tm
        self.genOmega = self.genOmegaS
        self.genTm = self.genE_d * self.genId + self.genE_q * self.genIq + (self.genX_q - self.genX_d) * self.genId * self.genIq
        # form the differential variable vector
        self.x = np.hstack((self.genE_q, self.genE_d, self.genDelta, self.genOmega, self.genEfd, self.genRf, self.genVr))
        # form the algebraic variable vector
        self.y = np.hstack((self.genId, self.genIq, self.busVm, self.busVa))

        return self.x, self.y

    def f(self, x, y):
        '''get the differential equations' result, including synchronise machines
        df = Ax + R(x, y) + Cu
        '''
        # detach the differential variable vector
        genE_q, genE_d, genDelta, genOmega, genEfd, genRf, genVr = np.split(x, 7)
        genId, genIq = np.split(y[:2*self.nGenerator], 2)
        busVm, busVa = np.split(y[2*self.nGenerator:], 2)
        genVm, genVa = busVm[self.genBus-1], busVa[self.genBus-1]
        # form the coefficient linear matrix in differential equations
        A11 = -np.diag(1/self.genT_d)
        A15 = -A11
        A22 = np.diag(-1/self.genT_q)
        A34 = np.eye(self.nGenerator)
        A44 = np.diag(-self.genD/self.genM)
        A55 = -np.diag(self.genKe/self.genTe)
        A57 = np.diag(1/self.genTe)
        A65 = np.diag(self.genKf / self.genTf ** 2)
        A66 = np.diag(-1/self.genTf)
        A75 = np.diag(-self.genKa*self.genKf/self.genTa/self.genTf)
        A76 = np.diag(self.genKa/self.genTa)
        A77 = np.diag(-1/self.genTa)
        # form the nonlinear matrix in differential equations
        R1 = -(self.genXd - self.genX_d)*genId/self.genT_d
        R2 = (self.genXq - self.genX_q)*genIq/self.genT_q
        R4 = -self.genOmegaS/2/self.genH*((genE_d*genId + genE_q*genIq) +
                                          (self.genX_q - self.genX_d)*genId*genIq)
        R5 = -self.genAx*np.exp(self.genBx*genEfd)*genEfd/self.genTe
        R7 = -self.genKa/self.genTa*genVm
        # form the common values matrix
        C31 = -1*np.eye(self.nGenerator)
        C41 = np.diag(self.genD/self.genM)
        C42 = np.diag(1/self.genM)
        C73 = np.diag(self.genKa/self.genTa)
        # form the differential variables vector
        dx1 = A11@genE_q + A15@genEfd + R1
        dx2 = A22@genE_d + R2
        dx3 = A34@genOmega + C31@self.genOmegaS
        dx4 = A44@genOmega + C41@self.genOmegaS + C42@self.genTm + R4
        dx5 = A55@genEfd + A57@genVr + R5
        dx6 = A65@genEfd + A66@genRf
        dx7 = A75@genEfd + A76@genRf + A77@genVr + C73@self.genVref + R7
        # assemble differentiation
        f = np.hstack((dx1, dx2, dx3, dx4, dx5, dx6, dx7))
        return f

    def g(self, x, y):
        '''includes two part.
        The 1st part is the stator algebraic equations,
        and the 2ed part is the network power flow equations
        0 = g(x, y)
        '''
        # detach the differential variable vector
        genE_q, genE_d, genDelta, genOmega, genEfd, genRf, genVr = np.split(x, 7)
        genId, genIq = np.split(y[:2 * self.nGenerator], 2)
        busVm, busVa = np.split(y[2 * self.nGenerator:], 2)
        genVm, genVa = busVm[self.genBus - 1], busVa[self.genBus - 1]
        # stator algebraic equations
        g1 = genE_d - genVm*np.sin(genDelta - genVa) - self.genRs*genId + self.genX_q*genIq
        g2 = genE_q - genVm*np.cos(genDelta - genVa) - self.genRs*genIq - self.genX_d*genId
        # network algebraic equations
        # load model is not included, which can be easily changed in the future
        vComplex = busVm*np.exp(1j*busVa)
        sComplex = vComplex*np.conj(self.Ybus@vComplex)
        g3 = (self.G2B@(genId*genVm*np.sin(genDelta - genVa) + genIq*genVm*np.cos(genDelta - genVa)) - self.busPd -
              sComplex.real)
        g4 = (self.G2B@(genId*genVm*np.cos(genDelta - genVa) - genIq*genVm*np.sin(genDelta - genVa)) - self.busQd -
              sComplex.imag)
        # assemble differentiation
        g = np.hstack((g1, g2, g3, g4))
        return g

    def fg(self, xy):
        '''assemble f and g'''
        x, y = xy[:self.nDimensionx], xy[self.nDimensionx:]
        return np.hstack((self.f(x, y), self.g(x, y)))

    def Jacobian(self, x, y):
        '''return the approximated jacobian matrix for Trapezoid, using the tiny differentiation method'''
        # difference disturbance
        disturbance = 10**(-9)
        # generate disturbed xy matrix corresponding to each variables
        xyMatrix = np.tile(np.hstack((x, y)), (self.nDimension, 1))
        xyMatrix_disturbance = xyMatrix + np.eye(self.nDimension) * disturbance
        # generate fg matrix and disturbed fg matrix corresponding to each variables in parallel
        fgMatrix = np.tile(self.fg(xyMatrix[0, :]), (self.nDimension, 1))
        fgMatrix_disturbance = [delayed(self.fg)(xyMatrix_disturbance[i, :]) for i in range(self.nDimension)]
        fgMatrix_disturbance = np.array(compute(*fgMatrix_disturbance))
        # compute dFx, dFy, dGx, dGy
        drivativefg = (fgMatrix_disturbance - fgMatrix)/disturbance
        # compute Jacobian matrix used for Trapezoid
        drivativefg[:self.nDimensionx, :self.nDimensionx] = drivativefg[:self.nDimensionx, :self.nDimensionx] - 2/self.deltaT*np.eye(self.nDimensionx)
        return drivativefg.transpose()

    def TransientSimulation(self, tEnd):
        '''simulate a transient simulation'''
        # disturbance only considering load now
        self.busPd[2:4] = self.busPd[2:4] + 3
        # set intermediate variables
        x, y = self.x, self.y
        # set a collection of trajectories
        self.xset, self.yset = list(), list()
        self.xset.append(x)
        self.yset.append(y)
        # start transient simulation
        for t in range(round(tEnd/self.deltaT)):
            f0, g = self.f(x, y), self.g(x, y)
            f = f0
            h = np.hstack((f + f0 -2/self.deltaT*(x - self.x), g))
            # start trapeziod iteration
            error = np.linalg.norm(h)
            error0 = error
            Jacobian = self.Jacobian(x, y)
            Jacobian = Jacobian + np.eye(self.nDimension) * 10 ** (-8)
            while error > 10**(-6):
                # if abs(np.linalg.det(Jacobian)) > 0.000001:
                #     deltaXY = np.linalg.solve(Jacobian, -h)
                # else:
                #     Jacobian = Jacobian + np.eye(self.nDimension)*10**(-8)
                #     deltaXY = np.linalg.solve(Jacobian, -h)
                deltaXY = np.linalg.solve(Jacobian, -h)
                lamdaRate = 1 # step size control
                while error0 >= error:
                    x_s, y_s = x + lamdaRate*deltaXY[:self.nDimensionx], y + lamdaRate*deltaXY[self.nDimensionx:]
                    f, g = self.f(x_s, y_s), self.g(x_s, y_s)
                    h = np.hstack((f + f0 - 2 / self.deltaT * (x_s - self.x), g))
                    error0 = np.linalg.norm(h)
                    lamdaRate = lamdaRate/2
                x, y = x_s, y_s
                error = error0
            self.x, self.y = x, y
            self.xset.append(x)
            self.yset.append(y)
        self.xset, self.yset = np.array(self.xset), np.array(self.yset)
        return self.xset, self.yset

if __name__ == '__main__':
    # run the transient dynamic
    main()
